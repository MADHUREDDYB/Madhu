import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet import dimensions
import datetime
import os


class QueryGenerator:

    def __init__(self, sttm_data,  output_folder_name, run_time):
        self.sttm_data = sttm_data
        self.output_folder_name = output_folder_name
        self.run_time = run_time

    def generate_queries(self,query_template_file,sub_folder):
        queries_template = openpyxl.load_workbook(query_template_file)
        queries_template_sheet = queries_template.active
        folder = self.output_folder_name+'\\Output Files'
        if not os.path.exists(folder):
            os.makedirs(folder)
        wb_output = Workbook()
        print(str(datetime.datetime.now()) + ' Query Generation.......')
        headers = ['Test Scenario ID', 'Test Case ID', 'Table Name','Column Name','Level', 'Test Case Name', 'Test Case Desc', 'Query','Exe Flag','Exe Status','Execpted Value','Actual Value','Result']
        tst_cs = 1
        # appending the data into list in the format of Src_cols,Tgt_cols,src_db,tgt_db,src_tb_name,tgt_tb_name,wf,mp
        for data in self.sttm_data:
            src_cols = data[0]
            tgt_cols = data[1]
            src_cols_Aud = [i for i in src_cols if ('AUD_' in i.upper())]
            tgt_cols_Aud = [i for i in tgt_cols if ('AUD_' in i.upper())]
            src_cols = [i for i in src_cols if not ('AUD_' in i.upper())]
            tgt_cols = [i for i in tgt_cols if not ('AUD_' in i.upper())]
            src_db = data[2]
            tgt_db = data[3]
            src_tb_nm = data[4]
            tgt_tb_nm = data[5]
            mpng = data[6]
            wf = data[7]
            trnsfrms = data[8]
            print(trnsfrms)
            trnsfrms_aud = [i for i in trnsfrms if ('AUD_' in i.upper())]
            trnsfrms = [i for i in trnsfrms if not ('AUD_' in i.upper())]
            trnsfrms = trnsfrms[:len(src_cols)]
            print(trnsfrms)
            src_fltr = data[9]
            src_join = data[10]
            tgt_fltr = data[11]
            tgt_join = data[12]
            btch_nm = data[14]
            ws1 = wb_output.create_sheet(title='%s.%s' % (tgt_db, tgt_tb_nm[len(tgt_tb_nm)-10:len(tgt_tb_nm)]))
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 10
            ws1.column_dimensions['C'].width = 15
            ws1.column_dimensions['D'].width = 15
            ws1.column_dimensions['E'].width = 10
            ws1.column_dimensions['F'].width = 20
            ws1.column_dimensions['G'].width = 30
            ws1.column_dimensions['H'].width = 50
            ws1.column_dimensions['I'].width = 10
            ws1.column_dimensions['J'].width = 10
            ws1.column_dimensions['K'].width = 10
            ws1.column_dimensions['L'].width = 10
            ws1.column_dimensions['M'].width = 10
            ws1.freeze_panes = 'B2'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A1:M1'
            ws1.append(headers)
            for m in range(1, len(headers)+1):
                ws1.cell(row=1, column=m).fill = PatternFill(start_color="575757", end_color="575757",
                                                             fill_type="solid")
                ws1.cell(row=1, column=m).font = Font(color='6FF242')
            row_num = 2  # sarting from 2 as first row has headers
            # for tst_desc, query in self.queries_data.items():
            for query_row in range(2,queries_template_sheet.max_row+1):
                ts_id = queries_template_sheet.cell(row=query_row,column=1).value
                level = queries_template_sheet.cell(row=query_row, column=5).value
                tc_name = queries_template_sheet.cell(row=query_row, column=6).value
                tc_desc = queries_template_sheet.cell(row=query_row, column=7).value
                query = queries_template_sheet.cell(row=query_row, column=8).value
                exe_flag = queries_template_sheet.cell(row=query_row, column=9).value
                exp_value = queries_template_sheet.cell(row=query_row, column=11).value
                if True:
                    if level.upper() == 'TABLE LEVEL':
                        query = self.DataReplace(query, src_cols, tgt_cols, src_db, tgt_db, src_tb_nm, tgt_tb_nm,trnsfrms,src_fltr,src_join,tgt_fltr,tgt_join,mpng,btch_nm, 'TABLELEVEL')
                        ws1.cell(row=row_num, column=1).value = ts_id
                        ws1.cell(row=row_num, column=2).value = "TC-" + str(row_num-1)
                        ws1.cell(row=row_num, column=3).value = tgt_db+'.'+tgt_tb_nm
                        ws1.cell(row=row_num, column=4).value = ','.join(tgt_cols)
                        ws1.cell(row=row_num, column=5).value = "Table Level"
                        ws1.cell(row=row_num, column=6).value = tc_name
                        ws1.cell(row=row_num, column=7).value = tc_desc
                        ws1.cell(row=row_num, column=8).value = query
                        ws1.cell(row=row_num, column=8).alignment = Alignment(wrap_text=True)
                        ws1.cell(row=row_num, column=8).alignment = Alignment(wrap_text=True)
                        cell_temp = ws1.cell(row=row_num, column=8).alignment
                        ws1.cell(row=row_num, column=9).value = exe_flag
                        ws1.cell(row=row_num, column=10).value = "Not Executed"
                        ws1.cell(row=row_num, column=11).value = exp_value
                        ws1.cell(row=row_num, column=12).value = ''
                        ws1.cell(row=row_num, column=13).value = ''
                        row_num += 1
                    else:
                        query_num = 0
                        for i in range(len(tgt_cols)):
                            query1 = self.DataReplace(query, src_cols, tgt_cols, src_db, tgt_db, src_tb_nm, tgt_tb_nm,trnsfrms, src_fltr,src_join,tgt_fltr,tgt_join,mpng,btch_nm,query_num)
                            ws1.cell(row=row_num, column=1).value = ts_id
                            ws1.cell(row=row_num, column=2).value = "TC-" + str(row_num-1)
                            ws1.cell(row=row_num, column=3).value = tgt_db + '.' + tgt_tb_nm
                            ws1.cell(row=row_num, column=4).value = tgt_cols[i]
                            ws1.cell(row=row_num, column=5).value = "Column Level"
                            ws1.cell(row=row_num, column=6).value = tc_name
                            ws1.cell(row=row_num, column=7).value = tc_desc
                            ws1.cell(row=row_num, column=8).value = query1
                            ws1.cell(row=row_num,column=8).alignment = Alignment(wrap_text=True)
                            ws1.cell(row=row_num, column=8).alignment = Alignment(wrap_text=True)
                            ws1.cell(row=row_num, column=9).value = exe_flag
                            ws1.cell(row=row_num, column=10).value = "Not Executed"
                            ws1.cell(row=row_num, column=11).value = exp_value
                            ws1.cell(row=row_num, column=12).value = ''
                            ws1.cell(row=row_num, column=13).value = ''
                            row_num += 1
                            tst_cs += 1
                            query_num+=1

                else:
                    pass
        wb_output.remove(wb_output['Sheet'])
        final_folder = folder + '\\Generated_ETL_Queries\\'+sub_folder
        if not os.path.exists(final_folder):
            os.makedirs(final_folder)
        wb_output.save(final_folder+'\\%s-%s_%s.xlsx'%(src_db,tgt_db,self.run_time))
        return final_folder

    def DataReplace(self, query, src_cols, tgt_cols, src_db, tgt_db, src_tb_nm, tgt_tb_nm,trnsfrmns, src_fltr,src_join,tgt_fltr,tgt_join,job_nm,btch_nm,flag):

        if src_fltr is None:
            src_fltr = '1 = 1'
        if src_join is None:
            src_join = '1 = 1'
        if tgt_fltr is None:
            tgt_fltr = '1 = 1'
        if tgt_join is None:
            tgt_join = '1 = 1'
        for i in range(len(src_cols)):
            if trnsfrmns[i] == "Direct Mapping":
                trnsfrmns[i] = src_cols[i]
        if flag == 'TABLELEVEL':
            query1 = query.replace('<TGT_COL_NM_ALL>', ','.join(tgt_cols)).replace('<SRC_COL_NM_ALL>',
                                                                                  ','.join(src_cols)).replace('<TGT_FLTR>', tgt_fltr).replace(
                '<TGT_JOIN>', tgt_join).replace('<DTA_TYP>', 'DATA_TYPE').replace('<SRC_TBL>', src_tb_nm).replace('<TRNS_ALL>',','.join(trnsfrmns)).replace('<SRC_FLTR>',src_fltr).replace('<SRC_JOIN>',src_join).replace('<SRC_DB>',src_db).replace('<TGT_TBL>',tgt_tb_nm).replace('<TGT_DB>',tgt_db)
        else:
            query1 = query.replace('<TGT_COL_NM>', tgt_cols[flag]).replace('<TGT_FLTR>', tgt_fltr).replace(
                '<TGT_JOIN>', tgt_join).replace('<DTA_TYP>', 'DATA_TYPE').replace('<SRC_TBL>', src_tb_nm).replace(
                '<SRC_COL_NM>', src_cols[flag]).replace('<SRC_FLTR>',src_fltr).replace('<SRC_JOIN>',src_join).replace('<TRNS>',trnsfrmns[flag]).replace('<SRC_DB>',src_db).replace('<TGT_TBL>',tgt_tb_nm).replace('<TGT_DB>',tgt_db)
        query1 = query1.replace('<JOB_NM>',job_nm).replace('<BATCH_NM>',btch_nm)
        return query1





if __name__=="__generate_queries__":
    generate_queries()