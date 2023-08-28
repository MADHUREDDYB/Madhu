import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
import os


class ReadConfigData:

    def __init__(self, src_sys_name, input_file_path,output_folder_nm, src_lake_file_name, lake_hist_file_name, lakehist_hub_file_name):
        self.src_sys_name = src_sys_name
        self.input_file_path = input_file_path
        self.src_lake_file_name = src_lake_file_name
        self.lake_hist_file_name = lake_hist_file_name
        self.lakehist_hub_file_name = lakehist_hub_file_name
        self.output_folder_nm = output_folder_nm


    def ReadExcelData(self):

        #Loading STTMs to Read Data
        src_lake_file_nm = '%s//Input Files//%s' % (self.input_file_path, self.src_lake_file_name)
        lake_hist_file_nm = '%s//Input Files//%s' % (self.input_file_path, self.lake_hist_file_name)
        hist_hub_file_nm = '%s//Input Files//%s' % (self.input_file_path, self.lakehist_hub_file_name)
        temp = "<b>EXECUTION STOPPED! : </b> File not found : "
        return_flag = False
        if not os.path.exists(src_lake_file_nm):
            return_flag = True
            temp += '<BR>' + str(src_lake_file_nm)
        if not os.path.exists(src_lake_file_nm):
            return_flag = True
            temp += '<BR>' + str(src_lake_file_nm)
        if not os.path.exists(lake_hist_file_nm):
            return_flag = True
            temp += '<BR>' + str(src_lake_file_nm)

        if return_flag == True:
             return temp

        wb_src_file = openpyxl.load_workbook(src_lake_file_nm)
        wb_lake_file = openpyxl.load_workbook(lake_hist_file_nm)
        wb_hist_file = openpyxl.load_workbook(hist_hub_file_nm)
        src_raw_data = []
        lake_data = []
        hist_data = []
        hub_data  = []
        #Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_src_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet in sheets:
            sheet = wb_src_file[sheet]
            src_file_name = sheet.cell(9, 1).value.replace('.csv', '')
            raw_post_path = sheet.cell(row=9, column=13).value
            lake_schema = sheet.cell(row=9, column=18).value
            lake_tab_name = sheet.cell(row=9, column=19).value
            src_raw_data.append([src_file_name, raw_post_path])
            lake_data.append([lake_schema, lake_tab_name])
        # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_lake_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet1 in sheets:
            sheet = wb_lake_file[sheet1]
            hist_work_flow = sheet.cell(row=4, column=2).value
            hist_schema = sheet.cell(row=7, column=2).value
            hist_tab_name = sheet.cell(row=11, column=6).value
            hist_cols = []
            for i in range(11, sheet.max_row):
                val = sheet.cell(row=i, column=7).value
                if val is not None and val != 'NA':
                    hist_cols.append(val)
                    if val == 'Aud_File_Nm':
                        break
            hist_data.append([hist_schema, hist_tab_name, hist_cols, hist_work_flow])
        # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_hist_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet1 in sheets:
                sheet = wb_hist_file[sheet1]
                hub_work_flow = sheet.cell(row=4, column=2).value
                hub_schema = sheet.cell(row=7, column=2).value
                hub_tab_name = sheet.cell(row=11, column=6).value
                hub_cols = []
                for i in range(11, sheet.max_row):
                    val = sheet.cell(row=i, column=7).value
                    if val is not None and val != 'NA':
                        hub_cols.append(val)
                        if val == 'Aud_Cur_Ind':
                            break
                hub_data.append([hub_schema, hub_tab_name, hub_cols, hub_work_flow])

        #Calling the Execute function in Begin class
        res = Begin.Execute(self.src_sys_name, self.src_sys_name, src_raw_data, lake_data, hist_data, hub_data, self.input_file_path,self.output_folder_nm)
        return "JIRA Test Case Generation is Completed and You can find the Output files at : %s" % res


class CreateSheet():
    tc_start_num = 0

    def __init__(self, sheet, src_sys_name, src_raw_data, lake_data, hist_data, hub_data, headers, input_file_path,
                 output_folder_name):
        self.sheet = sheet
        self.src_sys_name = src_sys_name
        self.src_raw_data = src_raw_data
        self.lake_data = lake_data
        self.hist_data = hist_data
        self.hub_data = hub_data
        self.headers = headers
        self.input_file_path = input_file_path
        self.output_folder_name = output_folder_name

    def srctoraw(self):
        tc_start_num = CreateSheet.tc_start_num
        wb_output = Workbook()
        print(str(datetime.datetime.now()) + ' Src to Raw JIRA preparation Started.......')
        sample_sheet = self.sheet["SRCTORAW"]
        max_rows = sample_sheet.max_row
        for i in range(len(self.src_raw_data)):
            entity = src_file_name = self.src_raw_data[i][0]
            raw_post_path = self.src_raw_data[i][1]
            test_summary = self.src_sys_name + '_TESTING_SRC_RAW_' + entity
            print(str(datetime.datetime.now()) + ' Creating %s Sheet' % entity)
            ws1 = wb_output.create_sheet(title=entity)
            ws1.append(self.headers)
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 40
            ws1.column_dimensions['D'].width = 30
            ws1.column_dimensions['F'].width = 40
            ws1.row_dimensions[1].height = 30
            ws1.freeze_panes = 'B2'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A1:R1'
            for m in range(1, 19):
                ws1.cell(row=1, column=m).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF",
                                                             fill_type="solid")
            for j in range(2, sample_sheet.max_row+1):
                ws1.row_dimensions[j].height = 30
                tc = str('TC' + str(tc_start_num))
                print(str(datetime.datetime.now()) + " Preparing TC NO: " + str(tc))
                for k in range(1, 19):
                    val = sample_sheet.cell(row=j, column=k).value
                    if val is not None:
                        val = str(val)
                        val = val.replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  src_file_name).replace(
                            '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                          raw_post_path)
                        ws1.cell(row=j, column=k).value = val
                tc_start_num += 1
        wb_output.remove(wb_output['Sheet'])
        print(str(datetime.datetime.now()) + " Src to Raw Completed............")
        print(str(datetime.datetime.now()) + " Creating the file at %s" % self.input_file_path)
        filename = (self.output_folder_name + '/%s_JIRA_SRCTORAW.xlsx' % self.src_sys_name)
        wb_output.save(filename)
        CreateSheet.tc_start_num = tc_start_num

    def rawtolake(self):
        print(str(datetime.datetime.now()) + ' Raw to Lake JIRA preparation Started.......')
        tc_start_num = CreateSheet.tc_start_num
        wb_output = Workbook()
        sample_sheet = self.sheet["RAWTOLAKE"]
        max_rows = sample_sheet.max_row
        for i in range(len(self.lake_data)):
            entity = self.src_raw_data[i][0]
            raw_post_path = self.src_raw_data[i][1]
            lake_schema = self.lake_data[i][0]
            lake_tab_name = self.lake_data[i][1]
            test_summary = self.src_sys_name + '_TESTING_RAW_LAKE_' + entity
            ws1 = wb_output.create_sheet(title=entity)
            ws1.append(self.headers)
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 40
            ws1.column_dimensions['D'].width = 30
            ws1.column_dimensions['F'].width = 40
            ws1.row_dimensions[1].height = 30
            ws1.freeze_panes = 'B2'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A1:R1'
            for m in range(1, 19):
                ws1.cell(row=1, column=m).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF",
                                                             fill_type="solid")
            for j in range(2, sample_sheet.max_row+1):
                ws1.row_dimensions[j].height = 30
                ws1.row_dimensions[j].height = 30
                tc = str('TC' + str(tc_start_num))
                print(str(datetime.datetime.now()) + " Preparing TC NO:  " + str(tc))
                for k in range(1, 19):
                    val = sample_sheet.cell(row=j, column=k).value
                    if val is not None:
                        val = str(val)
                        val = val.replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  entity).replace(
                            '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                               lake_tab_name).replace(
                            "<SOURCE_PATH>", raw_post_path)
                        ws1.cell(row=j, column=k).value = val
                tc_start_num += 1
        wb_output.remove(wb_output['Sheet'])
        print(str(datetime.datetime.now()) + " Raw to Lake Completed............")
        print(str(datetime.datetime.now()) + " Creating the file at %s" % (self.output_folder_name))
        filename = (self.output_folder_name + '/%s_JIRA_RawToLake.xlsx' % (self.src_sys_name))
        wb_output.save(filename)
        CreateSheet.tc_start_num = tc_start_num

    def laketohist(self):
        print(str(datetime.datetime.now()) + " Lake to Lake Hist Started............")
        tc_start_num = CreateSheet.tc_start_num
        wb_output = Workbook()
        sample_sheet = self.sheet["LAKETOHIST"]
        for i in range(len(self.hist_data)):
            lake_tab_name = entity = self.hist_data[i][1]
            hist_tab_name = self.hist_data[i][1]
            workflow = self.hist_data[i][3].replace('wf_', "")
            test_summary = self.src_sys_name + '_TESTING_LAKE_HIST_' + entity
            ws1 = wb_output.create_sheet(title=entity)
            ws1.append(self.headers)
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 40
            ws1.column_dimensions['D'].width = 30
            ws1.column_dimensions['F'].width = 40
            ws1.row_dimensions[1].height = 30
            ws1.freeze_panes = 'B2'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A1:R1'
            for m in range(1, 19):
                ws1.cell(row=1, column=m).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF",
                                                             fill_type="solid")
            start = 1
            for j in range(2, 10):
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[j].height = 30
                print("Preparing TC NO:" + str(tc))
                data = []
                for k in range(1, 19):
                    val = sample_sheet.cell(row=j, column=k).value
                    data.append(val)
                data[0] = start
                data[1] = test_summary
                data[2] = data[2].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  hist_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                data[3] = data[3].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  hist_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                data[5] = data[5].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  hist_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                tc_start_num += 1
                ws1.append(data)
                start += 1
            # cols_sheet = self.sheet["HIST-COLS"]
            data = []
            for k in range(1, 19):
                val = sample_sheet.cell(row=10, column=k).value
                data.append(val)
            print(data)
            cols = self.hist_data[i][2]
            for col in cols:
                ws1.row_dimensions[start].height = 30
                tc = str('TC' + str(tc_start_num))
                data[0] = start
                data[1] = test_summary
                data[2] = self.src_sys_name + '-BP-' + tc + '-LL-' + hist_tab_name + '-' + str(col) + '-' + 'S-T'
                tc_start_num += 1
                ws1.append(data)
                start += 1
            for col in cols:
                ws1.row_dimensions[start].height = 30
                tc = str('TC' + str(tc_start_num))
                data[0] = start
                data[1] = test_summary
                data[2] = self.src_sys_name + '-BP-' + tc + '-LL-' + hist_tab_name + '-' + str(col) + '-' + 'T-S'
                tc_start_num += 1
                ws1.append(data)
                start += 1
            for j in range(11, sample_sheet.max_row+1):
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[start].height = 30
                print(str(datetime.datetime.now()) + " Preparing TC NO: " + str(tc))
                data = []
                for k in range(1, 21):
                    val = sample_sheet.cell(row=j, column=k).value
                    data.append(val)
                data[0] = start
                data[1] = test_summary
                data[2] = data[2].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  lake_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                data[3] = data[3].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  hist_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                data[5] = data[5].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>',
                                                                                  hist_tab_name).replace(
                    '<TC_NAME>', tc).replace('<TARGET_PATH>', hist_tab_name).replace(
                    "<SOURCE_TABLE>", lake_tab_name).replace('<WORKFLOW>', workflow)
                tc_start_num += 1
                ws1.append(data)
                start += 1
        wb_output.remove(wb_output['Sheet'])
        print(str(datetime.datetime.now()) + " Lake to Lake Hist Completed............")
        print(str(datetime.datetime.now()) + " Creating the file at %s" % (self.output_folder_name))
        filename = (self.output_folder_name + '/%s_JIRA_LakeToLakeHist.xlsx' % (self.src_sys_name))
        wb_output.save(filename)
        CreateSheet.tc_start_num = tc_start_num

    def histtohub(self):
        print(str(datetime.datetime.now()) + " Lake Hist to Hub Started............")
        tc_start_num = CreateSheet.tc_start_num
        wb_output = Workbook()
        sample_sheet = self.sheet["LAKEHISTTOHUB"]
        for i in range(len(self.hub_data)):
            hub_tab_name = entity = self.hub_data[i][1]
            hist_tab_name = self.hist_data[i][1]
            workflow = self.hub_data[i][3].replace('wf_', "")
            test_summary = self.src_sys_name + '_TESTING_LAKE_HIST_' + entity
            ws1 = wb_output.create_sheet(title=entity)
            ws1.append(self.headers)
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 40
            ws1.column_dimensions['D'].width = 30
            ws1.column_dimensions['F'].width = 40
            ws1.row_dimensions[1].height = 30
            ws1.freeze_panes = 'B2'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A1:R1'
            for m in range(1, 19):
                ws1.cell(row=1, column=m).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF",
                                                             fill_type="solid")

            start = 1
            for j in range(2, 11):
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[j].height = 30
                print("Preparitng TC NO:" + str(tc))
                data = []
                for k in range(1, 20):
                    val = sample_sheet.cell(row=j, column=k).value
                    data.append(val)
                data[0] = start
                data[1] = test_summary
                data[2] = data[2].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", hist_tab_name).replace('<WORKFLOW>', workflow)
                data[3] = data[3].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", hist_tab_name).replace('<WORKFLOW>', workflow)
                data[5] = data[5].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", hist_tab_name).replace('<WORKFLOW>', workflow)
                tc_start_num += 1
                ws1.append(data)
                start += 1
            data = []
            for k in range(1, 19):
                val = sample_sheet.cell(row=12, column=k).value
                data.append(val)
            print(data)
            cols = self.hub_data[i][2]
            for col in cols:
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[start].height = 30
                data[0] = start
                data[1] = test_summary
                data[2] = self.src_sys_name + '-BP-' + tc + '-LH-' + hub_tab_name + '-' + str(col) + '-' + 'S-T'
                tc_start_num += 1
                ws1.append(data)
                start += 1
            for col in cols:
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[start].height = 30
                data[0] = start
                data[1] = test_summary
                data[2] = self.src_sys_name + '-BP-' + tc + '-LH-' + hub_tab_name + '-' + str(col) + '-' + 'T-S'
                tc_start_num += 1
                ws1.append(data)
                start += 1
            for j in range(12, sample_sheet.max_row+1):
                tc = str('TC' + str(tc_start_num))
                ws1.row_dimensions[start].height = 30
                print(str(datetime.datetime.now()) + " Preparing TC NO: " + str(tc))
                data = []
                for k in range(1, 21):
                    val = sample_sheet.cell(row=j, column=k).value
                    data.append(val)
                data[0] = start
                data[1] = test_summary
                data[2] = data[2].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", hist_tab_name).replace('<WORKFLOW>', workflow)
                data[3] = data[3].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", hist_tab_name).replace('<WORKFLOW>', workflow)
                data[5] = data[5].replace('<TEST_SUMMARY>', test_summary).replace('<TABLE_NAME>', hub_tab_name).replace(
                    '<TC_NAME>', tc).replace('<MAIN_NAME>', self.src_sys_name).replace('<TARGET_PATH>',
                                                                                  hub_tab_name).replace(
                    "<SOURCE_TABLE>", self.src_sys_name).replace('<WORKFLOW>', workflow)
                tc_start_num += 1
                ws1.append(data)
                start += 1
        wb_output.remove(wb_output['Sheet'])
        print(str(datetime.datetime.now()) + " Lake Hist to Hub Completed............")
        print(str(datetime.datetime.now()) + " Creating the file at %s" % (self.output_folder_name))
        filename = (self.output_folder_name + '/%s_JIRA_LakeHistToHub.xlsx' % (self.src_sys_name))
        wb_output.save(filename)
        CreateSheet.tc_start_num = tc_start_num


class Begin():

    def Execute(self, src_sys_name, src_raw_data, lake_data, hist_data, hub_data, input_file_path,output_folder_name):
        print(str(datetime.datetime.now()) + ' JIRA sheet Preparation Started..........')
        print(str(datetime.datetime.now()) + ' Reading the Input Sheet')
        wb_obj = openpyxl.load_workbook(input_file_path + '\\Pyfiles\\SAMPLES.xlsx')
        headers = ['External id', 'Test Summary', 'Test Name', 'Description', 'OrderId', 'Step', 'Test Data',
                   'Expected Result', 'Epic link', 'Linked issues', 'Issue Key [To add steps]', 'Issue Link Type',
                   'Issues Key To Link', 'Pre Condition', 'Post Condition', 'Test Method', 'Test Style', 'Test Level']
        CreateSheet.tc_start_num = 1
        folder = output_folder_name + '\\' + 'Output Files\\JIRA_Test_Case'
        if not os.path.exists(folder):
            os.makedirs(folder)
        output_folder_name = folder
        cs = CreateSheet(wb_obj, src_sys_name, src_raw_data, lake_data, hist_data, hub_data, headers, input_file_path,
                         output_folder_name)
        cs.srctoraw()
        cs.rawtolake()
        cs.laketohist()
        cs.histtohub()
        return folder


#################################################
######### JIRA Sheet Automation   ###############
#################################################


