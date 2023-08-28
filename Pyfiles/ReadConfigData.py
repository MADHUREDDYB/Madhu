import openpyxl
import os


class ReadConfigData:

    def __init__(self, src_sys_name, input_file_path, lake_hist_file_name, queries_data, exe_data):
        self.src_sys_name = src_sys_name
        self.input_file_path = input_file_path
        self.lake_hist_file_name = lake_hist_file_name
        self.queries_data = queries_data
        self.exe_data = exe_data

    def ReadSttmData(self, filename):

        # Loading STTMs to Read Data
        # if not os.path.exists(filename):
        #     return "<b>EXECUTION STOPPED! : </b> %s is not found " % filename
        wb_lake_file = openpyxl.load_workbook(filename)

        sttm_data = []
        # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_lake_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet1 in sheets:
            sheet = wb_lake_file[sheet1]
            tgt_mpng = sheet.cell(row=3, column=2).value
            btch_nm = sheet.cell(row=3, column=5).value
            tgt_workflow = sheet.cell(row=4, column=2).value
            src_db = sheet.cell(row=6, column=2).value
            tgt_db = sheet.cell(row=7, column=2).value
            src_tb_name = sheet.cell(row=11, column=1).value
            tgt_tb_name = sheet.cell(row=11, column=6).value
            src_fltr = sheet.cell(row=11, column=11).value
            src_join = sheet.cell(row=11, column=12).value
            temp_column = sheet.cell(row=11, column=13).value
            tgt_fltr = sheet.cell(row=11, column=14).value
            tgt_join = sheet.cell(row=11, column=15).value
            src_cols = []
            tgt_cols = []
            trnsfrmtns = []
            tgt_data_typs = []
            for i in range(11, sheet.max_row):
                temp_src_col = sheet.cell(row=i, column=2).value
                temp_tgt_col = sheet.cell(row=i, column=7).value
                temp_tgt_dttype = sheet.cell(row=i, column=8).value
                trnsfrmtns_temp = sheet.cell(row=i, column=5).value
                if temp_src_col is not None and temp_src_col.strip() != '':
                    src_cols.append(temp_src_col)
                if temp_tgt_col is not None and temp_tgt_col.strip() != '':
                    tgt_cols.append(temp_tgt_col)
                if trnsfrmtns_temp is not None and trnsfrmtns_temp.strip() != '':
                    trnsfrmtns.append(trnsfrmtns_temp)
                if temp_tgt_dttype is not None and temp_tgt_dttype.strip() != '':
                    tgt_data_typs.append(temp_tgt_dttype)

            # appending the data into list in the format of Src_cols,Tgt_cols,src_db,tgt_db,src_tb_name,tgt_tb_name,wf,mp,trnfromations,srcfltrand join, tgt flr an djoj
            sttm_data.append(
                [src_cols, tgt_cols, src_db, tgt_db, src_tb_name, tgt_tb_name, tgt_mpng, tgt_workflow, trnsfrmtns,
                 src_fltr, src_join, tgt_join, tgt_fltr, tgt_data_typs, btch_nm,temp_column])

        return sttm_data
        # Calling the Execute function in Begin class
        # Begin.Execute('', self.src_sys_name, sttm_data, self.queries_data, self.exe_data, input_file_path)

    def ReadLakeSttmData(self, filename):

        # Loading STTMs to Read Data
        # if not os.path.exists(filename):
        #     return "<b>EXECUTION STOPPED! : </b> %s is not found " % filename
        wb_lake_file = openpyxl.load_workbook(filename)
        sttm_data = []
        # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_lake_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet1 in sheets:
            sheet = wb_lake_file[sheet1]
            # tgt_mpng = sheet.cell(row=3, column=2).value
            # tgt_workflow = sheet.cell(row=4, column=2).value
            # src_db = sheet.cell(row=6, column=2).value
            tgt_db = sheet.cell(row=9, column=18).value
            # src_tb_name = sheet.cell(row=11, column=1).value
            tgt_tb_name = sheet.cell(row=9, column=19).value
            # src_fltr = sheet.cell(row=11, column=11).value
            # src_join = sheet.cell(row=11, column=12).value
            # tgt_fltr = sheet.cell(row=11, column=13).value
            # tgt_join = sheet.cell(row=11, column=14).value
            # tgt_data_typs = []
            tgt_cols = []
            # trnsfrmtns = []
            tgt_data_typs = []
            for i in range(9, sheet.max_row + 1):
                temp_tgt_col = sheet.cell(row=i, column=20).value
                temp_tgt_dttype = sheet.cell(row=i, column=21).value
                if temp_tgt_col is not None and temp_tgt_col != 'NA':
                    tgt_cols.append(temp_tgt_col)
                if temp_tgt_dttype is not None and temp_tgt_col != 'NA':
                    tgt_data_typs.append(temp_tgt_dttype)

            # appending the data into list in the format of Src_cols,Tgt_cols,src_db,tgt_db,src_tb_name,tgt_tb_name,wf,mp,trnfromations,srcfltrand join, tgt flr an djoj
            sttm_data.append([tgt_db, tgt_tb_name, tgt_cols, tgt_data_typs])

        return sttm_data
        # Calling the Execute function in Begin class
        # Begin.Execute('', self.src_sys_name, sttm_data, self.queries_data, self.exe_data, input_file_path)




if __name__ == "__ReadInputData__":
    ReadInputData()
if __name__ == "__ReadSTTMData__":
    ReadSttmData()
if __name__ == "__ReadConfigData__":
    ReadConfigData