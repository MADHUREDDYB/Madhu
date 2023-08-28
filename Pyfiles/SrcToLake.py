import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill , Font
import pandas as pd
import os
import datetime
import collections
from pymysql import *
import pandas.io.sql as sql
import csv




class ValidateSrcToLake():

    def __init__(self ,conn, output_folder_name, runtime):
        self.conn = conn
        self.output_folder_name = output_folder_name
        self.runtime = runtime

    def srctolake(self,src_file_name,raw_file_name,lake_tab_name,runtime):

        folder = self.output_folder_name + '\\' + 'Output Files'
        output_file_name = '' #Used for naming the generated output file name
        if not os.path.exists(folder):
            os.makedirs(folder)
        print(str(datetime.datetime.now()) + ' ' + 'Execution has Started....')
        raw = pd.read_csv(raw_file_name)
        if src_file_name != False:
            src = pd.read_csv(src_file_name)
            headers = list(src.columns.values)
            flag = 'Src-Raw'
            output_file_name = str(src_file_name).split("\\")
            output_file_name.reverse()
            output_file_name = output_file_name[0].replace('.csv','') + '-Raw'
        else:
            src = raw
            headers = list(raw.columns.values)
        if  lake_tab_name != False:
            lake = sql.read_sql('select * from %s' % lake_tab_name, self.conn)
            flag = 'Raw-Lake'
            output_file_name = str(raw_file_name).split("\\")
            output_file_name.reverse()
            output_file_name = output_file_name[0].replace('.csv', '') + '-Lake'
        else:
            lake = raw
        if src_file_name and lake_tab_name:
            flag = 'Src-Lake'
            output_file_name = str(src_file_name).split("\\")
            output_file_name.reverse()
            output_file_name = output_file_name[0].replace('.csv', '') + '-Lake'
        final_folder = folder + '\\Src-Lake_Validations\\'+flag
        if not os.path.exists(final_folder):
            os.makedirs(final_folder)
        with pd.ExcelWriter(final_folder+'\\%s_%s.xlsx' %(output_file_name,self.runtime)) as writer:
            src.to_excel(writer, sheet_name='Src', index=False)
            raw.to_excel(writer, sheet_name='Raw', index=False)
            lake.to_excel(writer, sheet_name='Lake', index=False)
        print(str(datetime.datetime.now()) + ' ' + 'Creating the Excel Sheet')
        src_obj = openpyxl.load_workbook(final_folder+'\\%s_%s.xlsx' %(output_file_name,self.runtime))
        os.remove(final_folder+'\\%s_%s.xlsx' %(output_file_name,self.runtime))
        src_sheet = src_obj['Src']
        raw_sheet = src_obj['Raw']
        lake_sheet = src_obj['Lake']
        print(str(datetime.datetime.now()) + ' ' + 'Creating new sheet for Data validation')
        output_validation = src_obj.create_sheet(title="Src_Raw_Validation")
        output_validation2 = src_obj.create_sheet(title="Raw_Lake_Validation")
        if src_file_name != False:
            src_max_rows = src_sheet.max_row
            src_max_cols = src_sheet.max_column
            raw_max_rows = raw_sheet.max_row
            raw_max_cols = raw_sheet.max_column
        else:
            src_max_rows = raw_sheet.max_row
            src_max_cols = raw_sheet.max_column
            raw_max_rows = lake_sheet.max_row
            raw_max_cols = lake_sheet.max_column
        if src_max_rows < raw_max_rows:
            src_max_rows = raw_max_rows
        src_raw_mismatched_Data = []
        src_raw_error_rcd_cnt = []
        raw_lake_mismatched_Data = []
        raw_lake_error_rcd_cnt = []
        print(str(datetime.datetime.now()) + ' ' + 'Src and Raw Data Vadlidation Started........')
        for row in range(2, src_max_rows + 1):
            for col in range(1, src_max_cols + 1):
                src_value = src_sheet.cell(row, col).value
                raw_value = raw_sheet.cell(row, col).value
                lake_value = lake_sheet.cell(row, col).value
                output_validation.cell(row, col).value = src_value
                output_validation.cell(row, col).value = raw_value
                output_validation2.cell(row, col).value = raw_value
                # checking src and Raw values
                if src_value == raw_value:
                    output_validation.cell(row, col).value = 'True'
                    output_validation.cell(row=row, column=col).fill = PatternFill(start_color="8DFE90",
                                                                                   end_color="8DFE90",
                                                                                   fill_type="solid")
                else:
                    string = ""
                    j = col
                    while j > 0:
                        j, remainder = divmod(j - 1, 26)
                        string = chr(65 + remainder) + string
                    src_raw_error_rcd_cnt.append(row)
                    src_raw_mismatched_Data.append(
                        [string + str(row), src_value, raw_value, 'Mismatch Found'])
                    output_validation.cell(row=row, column=col).value = 'False'
                    output_validation.cell(row=row, column=col).fill = PatternFill(start_color="FE635E",
                                                                                   end_color="FE635E",
                                                                                   fill_type="solid")
                # raw and lake check
                if raw_value == lake_value:
                    output_validation2.cell(row, col).value = 'True'
                    output_validation2.cell(row=row, column=col).fill = PatternFill(start_color="8DFE90",
                                                                                    end_color="8DFE90",
                                                                                    fill_type="solid")
                else:
                    string = ""
                    j = col
                    while j > 0:
                        j, remainder = divmod(j - 1, 26)
                        string = chr(65 + remainder) + string
                    raw_lake_error_rcd_cnt.append(row)
                    raw_lake_mismatched_Data.append(
                        [string + str(row), raw_value, lake_value, 'Mismatch Found'])
                    output_validation2.cell(row=row, column=col).value = 'False'
                    output_validation2.cell(row=row, column=col).fill = PatternFill(start_color="FE635E",
                                                                                    end_color="FE635E",
                                                                                    fill_type="solid")
        print(str(datetime.datetime.now()) + ' ' + 'Validation Completed........')
        counter = collections.Counter(src_raw_error_rcd_cnt)
        counter2 = collections.Counter(raw_lake_error_rcd_cnt)
        src_raw_error_rcd_cnt = len(list(counter.keys()))
        raw_lake_error_rcd_cnt = len(list(counter2.keys()))
        print(str(datetime.datetime.now()) + ' Mismatcheds Found in ' + str(src_raw_error_rcd_cnt) + ' Records')
        if  src_file_name != False:
            src_mismatched_Data_Sheet = src_obj.create_sheet(title="Src-Raw-Mismatched Data")
            src_mismatched_Data_Sheet.append(['Record Number', 'Src Data', 'Raw Data'])
            src_mismatched_Data_Sheet.column_dimensions['A'].width = 10
            src_mismatched_Data_Sheet.column_dimensions['B'].width = 40
            src_mismatched_Data_Sheet.column_dimensions['C'].width = 40
            src_mismatched_Data_Sheet.column_dimensions['D'].width = 30
            src_mismatched_Data_Sheet.freeze_panes = 'B2'
            src_mismatched_Data_Sheet.sheet_view.zoomScale = 70
            src_mismatched_Data_Sheet.auto_filter.ref = 'A1:M1'
        if lake_tab_name != False:
            lake_mismatched_Data_Sheet = src_obj.create_sheet(title="Raw-Lake-Mismatched Data")
            lake_mismatched_Data_Sheet.append(['Record Number', 'Raw Data', 'Lake Data'])
            lake_mismatched_Data_Sheet.column_dimensions['A'].width = 10
            lake_mismatched_Data_Sheet.column_dimensions['B'].width = 40
            lake_mismatched_Data_Sheet.column_dimensions['C'].width = 40
            lake_mismatched_Data_Sheet.column_dimensions['D'].width = 30
            lake_mismatched_Data_Sheet.freeze_panes = 'B2'
            lake_mismatched_Data_Sheet.sheet_view.zoomScale = 70
            lake_mismatched_Data_Sheet.auto_filter.ref = 'A1:M1'

        if src_file_name != False:
            if src_raw_mismatched_Data:
                src_raw_mismatched_Data.sort(key=lambda x: x[0])
                src_raw_mismatched_Data.append(
                    ["", '', '', 'Mismatche(s) Found in %s records ' % (str(src_raw_error_rcd_cnt))])
                print(str(datetime.datetime.now()) + ' ' + 'Adding Mismatched Data to Sheet')
                for i in range(len(src_raw_mismatched_Data)):
                    src_mismatched_Data_Sheet.sheet_properties.tabColor = 'FE635E'
                    src_mismatched_Data_Sheet.append(src_raw_mismatched_Data[i])
                    src_mismatched_Data_Sheet.cell(row=i + 2, column=4).fill = PatternFill(start_color="FE635E",
                                                                                           end_color="FE635E",
                                                                                           fill_type="solid")
            else:
                src_mismatched_Data_Sheet.title = 'SR-Summary'
                src_mismatched_Data_Sheet.sheet_properties.tabColor = '75FB17'
                src_mismatched_Data_Sheet.cell(row=2, column=4).value = 'Validation Is Successful'
                src_mismatched_Data_Sheet.cell(row=2, column=4).fill = PatternFill(start_color="75FB17",
                                                                                   end_color="75FB17",
                                                                                   fill_type="solid")

            # adding headers adn styles to headers
            print(str(datetime.datetime.now()) + ' ' + 'Styling the Excel with colors........')
            for i in range(1, 4):
                src_mismatched_Data_Sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                                               fill_type="solid")
                src_mismatched_Data_Sheet.cell(row=1, column=i).font = Font(color='6FF242')
            for i in range(1, len(headers) + 1):
                output_validation.cell(row=1, column=i).value = headers[i - 1]
                string = ""
                j = i
                while j > 0:
                    j, remainder = divmod(j - 1, 26)
                    string = chr(65 + remainder) + string
                src_sheet.column_dimensions[string].width = 30
                raw_sheet.column_dimensions[string].width = 30
                output_validation.column_dimensions[string].width = 30
                raw_sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                               fill_type="solid")
                raw_sheet.cell(row=1, column=i).font = Font(color='6FF242')
                src_sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                               fill_type="solid")
                src_sheet.cell(row=1, column=i).font = Font(color='6FF242')
                output_validation.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                                       fill_type="solid")
                output_validation.cell(row=1, column=i).font = Font(color='6FF242')

        #############################LKW DATA
        if lake_tab_name != False:
            if raw_lake_mismatched_Data:
                raw_lake_mismatched_Data.sort(key=lambda x: x[0])
                raw_lake_mismatched_Data.append(
                    ["", '', '', 'Mismatche(s) Found in %s records ' % (str(raw_lake_error_rcd_cnt))])
                print(str(datetime.datetime.now()) + ' ' + 'Adding Mismatched Data to Sheet')
                for i in range(len(raw_lake_mismatched_Data)):
                    lake_mismatched_Data_Sheet.sheet_properties.tabColor = 'FE635E'
                    lake_mismatched_Data_Sheet.append(raw_lake_mismatched_Data[i])
                    lake_mismatched_Data_Sheet.cell(row=i + 2, column=4).fill = PatternFill(start_color="FE635E",
                                                                                            end_color="FE635E",
                                                                                            fill_type="solid")
            else:
                lake_mismatched_Data_Sheet.title = 'RL-Summary'
                lake_mismatched_Data_Sheet.sheet_properties.tabColor = '75FB17'
                lake_mismatched_Data_Sheet.cell(row=2, column=4).value = 'Validation Is Successful'
                lake_mismatched_Data_Sheet.cell(row=2, column=4).fill = PatternFill(start_color="75FB17",
                                                                                    end_color="75FB17",
                                                                                    fill_type="solid")

            # adding headers adn styles to headers
            print(str(datetime.datetime.now()) + ' ' + 'Styling the Excel with colors........')
            for i in range(1, 4):
                lake_mismatched_Data_Sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757",
                                                                                    end_color="575757",
                                                                                    fill_type="solid")
                lake_mismatched_Data_Sheet.cell(row=1, column=i).font = Font(color='6FF242')
            for i in range(1, len(headers) + 1):
                output_validation2.cell(row=1, column=i).value = headers[i - 1]
                string = ""
                j = i
                while j > 0:
                    j, remainder = divmod(j - 1, 26)
                    string = chr(65 + remainder) + string
                raw_sheet.column_dimensions[string].width = 30
                lake_sheet.column_dimensions[string].width = 30
                output_validation2.column_dimensions[string].width = 30
                lake_sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                                    fill_type="solid")
                lake_sheet.cell(row=1, column=i).font = Font(color='6FF242')
                raw_sheet.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                                   fill_type="solid")
                raw_sheet.cell(row=1, column=i).font = Font(color='6FF242')
                output_validation2.cell(row=1, column=i).fill = PatternFill(start_color="575757", end_color="575757",
                                                                            fill_type="solid")
                output_validation2.cell(row=1, column=i).font = Font(color='6FF242')

        src_sheet.freeze_panes = 'B2'
        src_sheet.sheet_view.zoomScale = 70
        src_sheet.auto_filter.ref = 'A1:M1'
        raw_sheet.freeze_panes = 'B2'
        raw_sheet.sheet_view.zoomScale = 70
        raw_sheet.auto_filter.ref = 'A1:M1'
        lake_sheet.freeze_panes = 'B2'
        lake_sheet.sheet_view.zoomScale = 70
        lake_sheet.auto_filter.ref = 'A1:M1'
        output_validation.freeze_panes = 'B2'
        output_validation.sheet_view.zoomScale = 70
        output_validation.auto_filter.ref = 'A1:M1'
        output_validation2.freeze_panes = 'B2'
        output_validation2.sheet_view.zoomScale = 70
        output_validation2.auto_filter.ref = 'A1:M1'
        if src_file_name == False:
            src_obj.remove(src_sheet)
            # src_obj.remove(src_mismatched_Data_Sheet)
            src_obj.remove(output_validation)
        elif lake_tab_name == False:
            src_obj.remove(lake_sheet)
            # src_obj.remove(lake_mismatched_Data_Sheet)
            src_obj.remove(output_validation2)

        src_obj.save(final_folder+'\\%s_%s.xlsx' %(output_file_name,runtime))
        return final_folder




if __name__ == "__ValidateSrcToLake__":
    ValidateSrcToLake
if __name__ == "__srctolake__":
    srctolake()