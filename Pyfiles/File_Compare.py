import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font
import pandas as pd
import os
import datetime
import collections
import numpy as np

class FileCompare():

    def __init__(self, runtime,input_file_path,output_file_path, src_sys_name,primary_file_name,secondary_file_name):
        self.runtime = runtime
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.src_sys_name = src_sys_name
        self.primary_file_name = primary_file_name
        self.secondary_file_name = secondary_file_name

    def compare_files(self):
        print(str(datetime.datetime.now())+' '+ 'Executing the File Comparision Function')
        folder = self.output_file_path+'\\Output Files\\File Compare'
        if not os.path.exists(folder):
            os.makedirs(folder)
        primary_file_name = self.input_file_path+'\\Input Files\\'+self.primary_file_name
        secondary_file_name =self.input_file_path+'\\Input Files\\'+ self.secondary_file_name
        primary_wb = openpyxl.load_workbook(primary_file_name)
        seconday_wb = openpyxl.load_workbook(secondary_file_name)
        primary_wb_sheets  = primary_wb.sheetnames
        seconday_wb_sheets = seconday_wb.sheetnames
        print(primary_wb_sheets)
        print(seconday_wb_sheets)
        sheets_len = 0
        if len(primary_wb_sheets) > len(seconday_wb_sheets):
            sheets_len = len(primary_wb_sheets)
            sheet_names = primary_wb_sheets
        else:
            sheets_len = len(seconday_wb_sheets)
            sheet_names = seconday_wb_sheets
        primary_wb_sheets += ['Empty'] * (sheets_len - len(primary_wb_sheets))
        seconday_wb_sheets += ['Empty'] * (sheets_len - len(seconday_wb_sheets))
        compare_wb = Workbook()
        total_data = []
        for i in range(sheets_len):
            mismatched_data = []
            primary_sheet = primary_wb_sheets[i]
            secondary_sheet = seconday_wb_sheets[i]
            print(primary_wb_sheets)
            print(seconday_wb_sheets)
            if primary_sheet == 'Empty':
                mismatched_data.append(["Data Not found",'Sheet Deleted',' '])
            elif secondary_sheet == 'Empty':
                mismatched_data.append(["Data Not found", ' ', 'Sheet Deleted'])
            else:
                primary_sheet = primary_wb[primary_wb_sheets[i]]
                secondary_sheet = seconday_wb[seconday_wb_sheets[i]]
                max_rw = max(primary_sheet.max_row, secondary_sheet.max_row)
                max_cl = max(primary_sheet.max_column, secondary_sheet.max_column)
                for row_num in range(1, max_rw + 1):
                    for col_num in range(1, max_cl + 1):
                        primary_sheet_val = primary_sheet.cell(row=row_num, column=col_num).value
                        seconday_sheet_val = secondary_sheet.cell(row=row_num, column=col_num).value
                        if primary_sheet_val != seconday_sheet_val:
                            string = ""
                            j = col_num
                            while j > 0:
                                j, remainder = divmod(j - 1, 26)
                                string = chr(65 + remainder) + string
                            mismatched_data.append([string + str(row_num), primary_sheet_val, seconday_sheet_val])
            total_data.append(mismatched_data)
        ws1 = compare_wb.create_sheet(title='Summary')
        row_num = 1
        end_col = 0
        for mismatched_data in total_data:
            ws1.cell(row=1, column=end_col + 1).value = "Sheet" + str(row_num)
            ws1.cell(row=2,column=end_col+1).value = 'Cell'
            ws1.cell(row=2, column=end_col + 2).value = 'Primary Sheet Value'
            ws1.cell(row=2, column=end_col + 3).value = 'Secondary Sheet Value'
            string = ""
            j = end_col+1
            while j > 0:
                j, remainder = divmod(j - 1, 26)
                string = chr(65 + remainder) + string
            ws1.column_dimensions[string].width = 25
            string = ""
            j = end_col+2
            while j > 0:
                j, remainder = divmod(j - 1, 26)
                string = chr(65 + remainder) + string
            ws1.column_dimensions[string].width = 20
            string = ""
            j = end_col+3
            while j > 0:
                j, remainder = divmod(j - 1, 26)
                string = chr(65 + remainder) + string
            ws1.column_dimensions[string].width = 25
            ws1.merge_cells(start_row=1, start_column=end_col+1, end_row=1, end_column=end_col + 3)
            for i  in range(len(mismatched_data)):
                data = mismatched_data[i]
                ws1.cell(row=i+3, column=end_col+1).value = data[0]
                ws1.cell(row=i+3, column=end_col+2).value = data[1]
                ws1.cell(row=i+3, column=end_col+3).value = data[2]
            end_col += 3
            row_num += 1
        print(str(datetime.datetime.now()) + ' ' + 'Comparsion file Created at ' + self.output_file_path + '//' + 'filecompare.xlsx')
        compare_wb.remove(compare_wb['Sheet'])
        ws1.freeze_panes = 'A2'
        j = ws1.max_column
        print(j)
        while j > 0:
            j, remainder = divmod(j - 1, 26)
            string = chr(65 + remainder) + string
        print(string)
        ws1.sheet_view.zoomScale = 70
        ws1.auto_filter.ref = 'A2:%s2' % string[0]
        for m in range(1, ws1.max_column):
            ws1.cell(row=1, column=m).fill = PatternFill(start_color="575757", end_color="575757",
                                                         fill_type="solid")
            ws1.cell(row=1, column=m).font = Font(color='6FF242')
        compare_wb.save(folder + '\\' + 'File_Compare_%s.xlsx' %self.runtime)
        return 'File Comparsion is Successfull and validation file is placed at %s' %(folder)

