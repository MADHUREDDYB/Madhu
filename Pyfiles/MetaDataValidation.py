import copy
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet import dimensions
import datetime
import os
from pymysql import *
import pandas.io.sql as sql
from collections import OrderedDict


class Metadatavalidation:

    def __init__(self, sttm_data, output_folder_nm, conn, run_time):
        self.sttm_data = sttm_data
        self.output_folder_nm = output_folder_nm
        self.conn = conn
        self.run_time = run_time

    def validate_metadata(self, flag):
        folder = self.output_folder_nm + '\\' + 'Output Files'
        if not os.path.exists(folder):
            os.makedirs(folder)
        wb_output = Workbook()
        summary = []
        for data in self.sttm_data:
            summary_flag = True
            if flag != 'Lake':
                tgt_cols = data[1]
                tgt_db = data[3]
                src_tb = data[4]
                tgt_tb_nm = data[5]
                tgt_data_types = data[13]
                tgt_res = {tgt_cols[i]: tgt_data_types[i] for i in range(len(tgt_data_types))}
            else:
                src_tb = data[1]
                tgt_db = data[0]
                tgt_tb_nm = data[1]
                tgt_cols = data[2]
                tgt_data_types = data[3]
                tgt_res = {tgt_cols[i]: tgt_data_types[i] for i in range(len(tgt_data_types))}
            query = "Desc %s.%s" % (tgt_db, tgt_tb_nm)
            db_df = sql.read_sql(query, self.conn)
            db_cols = db_df.col_name.tolist()
            db_data_types = db_df.data_type.tolist()
            # db_res = {db_cols[i]: db_data_types[i] for i in range(len(db_data_types))}
            # tgt_res= OrderedDict(sorted(tgt_res.items()))
            # db_res= OrderedDict(sorted(db_res.items()))
            # tgt_cols = list(tgt_res.keys())
            # tgt_data_types = list(tgt_res.values())
            # db_cols = list(db_res.keys())
            # db_data_types = list(db_res.values())
            ws1 = wb_output.create_sheet(title=src_tb)
            ws1.cell(row=1, column=1).value = "STTM"
            ws1.cell(row=1, column=3).value = 'DB'
            ws1.cell(row=1, column=5).value = 'VALIDATION'
            headers = ['Column Name', 'Data Type', 'Column Name', 'Data Type', 'Column Name', 'Data Type']
            ws1.append(headers)
            ws1.merge_cells('A1:B1')
            ws1.merge_cells('C1:D1')
            ws1.merge_cells('E1:F1')
            ws1.column_dimensions['A'].width = 30
            ws1.column_dimensions['B'].width = 30
            ws1.column_dimensions['C'].width = 30
            ws1.column_dimensions['D'].width = 30
            ws1.column_dimensions['E'].width = 30
            ws1.column_dimensions['F'].width = 30
            ws1.cell(1, 1).alignment = Alignment(horizontal='center')
            ws1.cell(1, 3).alignment = Alignment(horizontal='center')
            ws1.cell(1, 5).alignment = Alignment(horizontal='center')

            for m in range(1, len(headers) + 1):
                ws1.cell(row=2, column=m).fill = PatternFill(start_color="4294F2", end_color="4294F2",
                                                             fill_type="solid")
                ws1.cell(row=2, column=m).font = Font(color='FBFBFB')
            ws1.cell(row=1, column=1).fill = PatternFill(start_color="575757", end_color="575757",
                                                         fill_type="solid")
            ws1.cell(row=1, column=1).font = Font(color='6FF242')
            ws1.cell(row=1, column=3).fill = PatternFill(start_color="575757", end_color="575757",
                                                         fill_type="solid")
            ws1.cell(row=1, column=3).font = Font(color='6FF242')
            ws1.cell(row=1, column=5).fill = PatternFill(start_color="575757", end_color="575757",
                                                         fill_type="solid")  # 6FF242
            ws1.cell(row=1, column=5).font = Font(color='6FF242')
            ws1.freeze_panes = 'A3'
            ws1.sheet_view.zoomScale = 70
            ws1.auto_filter.ref = 'A2:F2'
            for i in range(1, len(tgt_cols) + 1):
                ws1.cell(row=i + 2, column=1).value = tgt_cols[i - 1]
                ws1.cell(row=i + 2, column=3, value='content').value = str(db_cols[i - 1])
                ws1.cell(row=i + 2, column=2).value = tgt_data_types[i - 1]
                ws1.cell(row=i + 2, column=4, value='content').value = str(db_data_types[i - 1])
                if tgt_cols[i - 1].upper().strip() == str(db_cols[i - 1]).upper().strip():
                    ws1.cell(row=i + 2, column=5).value = "TRUE"
                else:
                    summary_flag = False
                    ws1.cell(row=i + 2, column=5).value = "FALSE"
                if tgt_data_types[i - 1].upper().strip() == str(db_data_types[i - 1]).upper().strip():
                    ws1.cell(row=i + 2, column=6).value = "TRUE"
                else:
                    summary_flag = False
                    ws1.cell(row=i + 2, column=6).value = "FALSE"
            if not summary_flag:
                summary.append(['%s.%s' % (tgt_db, tgt_tb_nm), 'Fail'])
        final_folder = folder + '\\MetadaValidations\\%s' % flag
        if not os.path.exists(final_folder):
            os.makedirs(final_folder)
        summary_sheet = wb_output.create_sheet(title='Summary',index=1)
        wb_output.active = summary_sheet
        summary_sheet.freeze_panes = 'A3'
        summary_sheet.sheet_view.zoomScale = 70
        summary_sheet.auto_filter.ref = 'A1:B1'
        summary_sheet.column_dimensions['A'].width = 100
        summary_sheet.column_dimensions['B'].width = 30
        summary_headers = ['Table Name', 'Status']
        summary_sheet.append(summary_headers)
        summary_sheet.cell(row=1, column=1).font = Font(color='6FF242')
        summary_sheet.cell(row=1, column=1).fill = PatternFill(start_color="575757", end_color="575757",
                                                               fill_type="solid")
        summary_sheet.cell(row=1, column=2).font = Font(color='6FF242')
        summary_sheet.cell(row=1, column=2).fill = PatternFill(start_color="575757", end_color="575757",
                                                               fill_type="solid")
        if not summary:
            summary_sheet.append(['', '', 'Validation Successful'])
        else:
            for i in summary:
                summary_sheet.append(i)
        wb_output.remove(wb_output['Sheet'])
        wb_output.save(final_folder + '\\%s_%s.xlsx' % (tgt_tb_nm, self.run_time))
        return final_folder


#
if __name__ == "__Metadatavalidation__":
    Metadatavalidation

if __name__ == "__validate_metadata__":
    validate_metadata()
