import openpyxl
import os
from openpyxl import Workbook
import datetime

class ValidateSTTM:

    def __init__(self,runtime, input_file_path, output_file_path,src_sys_name):
        self.runtime = runtime
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.src_sys_name = src_sys_name

    def validate_lake_mart_sttm(self, filename):

        folder = self.input_file_path+'\\Output Files\\'+'STTM Validations'
        if not os.path.exists(folder):
            os.makedirs(folder)
        filename = self.input_file_path+'\\Input Files\\'+ filename
        # Loading STTMs to Read Data
        # if not os.path.exists(filename):
        #     return "<b>EXECUTION STOPPED! : </b> %s is not found " % filename
        template = self.input_file_path + '\\Pyfiles\\STTM_Lake_Mart_Template.xlsx'
        wb_output = Workbook()
        template_sheet = openpyxl.load_workbook(template)
        template_sheet = template_sheet['Lake-Mart-Template']
        wb_lake_file = openpyxl.load_workbook(filename)

        validated_data = []
        # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
        sheets = list(wb_lake_file.sheetnames)
        if sheets[0].upper() == "COVER":
            del sheets[0]
        for sheet1 in sheets:
            sheet_validated_data = {}
            sheet = wb_lake_file[sheet1]
            if sheet.cell(row=3, column=2).value != template_sheet.cell(row=3, column=2).value and sheet.cell(row=3, column=3) is None:
                sheet_validated_data['Mapping Name'] = 'Not Provided'
            if sheet.cell(row=4, column=2).value != template_sheet.cell(row=3, column=2).value and sheet.cell(row=3, column=3) is None:
                sheet_validated_data['WorkFlow Name'] = 'Not Provided'
            if sheet.cell(row=5, column=2).value != template_sheet.cell(row=3, column=2).value and sheet.cell(row=3, column=3) is None:
                sheet_validated_data['LoadStrategy'] = 'Not Provided'
            if sheet.cell(row=6, column=2).value != template_sheet.cell(row=3, column=2).value and sheet.cell(row=3, column=3) is None:
                sheet_validated_data['Source DB'] = 'Not Provided'
            if sheet.cell(row=7, column=2).value != template_sheet.cell(row=3, column=2).value and sheet.cell(row=3, column=3) is None:
                sheet_validated_data['Target DB'] = 'Not Provided'
            for i in range(9,11):
                for j in range(1,template_sheet.max_row):
                    if sheet.cell(row=i, column=j).value != template_sheet.cell(row=i, column=j).value:
                        sheet_validated_data['Headers-CELL(%s,%s)' %(i,j)] = 'Headers Not Matched With Template'
            for i in range(11,sheet.max_row):
                for j in range(1,9):
                    if sheet.cell(row=i, column=j).value is None:
                        sheet_validated_data['Data-CELL(%s,%s)' %(i,j)] = 'Mandatory Data should not be Empty.Give NA if Not applicable'
            if bool(sheet_validated_data):
                # Dictionary is Empty
                validated_data.append(sheet_validated_data)
                temp_sheet = wb_output.create_sheet(title=sheet1)
                temp_sheet.append(['Cell Number','Summary'])
                row = 2
                for key,value in sheet_validated_data.items():
                    #If you need any colring add here
                    temp_sheet.cell(row=row,column=1).value = key
                    temp_sheet.cell(row=row, column=2).value = value
                    temp_sheet.column_dimensions['A'].width = 50
                    temp_sheet.column_dimensions['B'].width = 100
                    row += 1
        print(validated_data)
        run_time = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S").replace(" ", '-').replace(':',
                                                                                                   '').replace(
            '/', '')
        if validated_data:
            wb_output.remove(wb_output['Sheet'])
            wb_output.save(folder+'\\'+'%s_STTMValidation.xlsx' %run_time)
            return 'STTM Validation failed for %s sheet in %s file' %(filename,sheet1)
        return ''
        # Calling the Execute function in Begin class
        # Begin.Execute('', self.src_sys_name, sttm_data, self.queries_data, self.exe_data, input_file_path)

    # def ReadLakeSttmData(self, filename):
    #
    #     # Loading STTMs to Read Data
    #     # if not os.path.exists(filename):
    #     #     return "<b>EXECUTION STOPPED! : </b> %s is not found " % filename
    #     wb_lake_file = openpyxl.load_workbook(filename)
    #     sttm_data = []
    #     # Reading Src file Sheet Names into a List to remove The Sheet named "Cover"
    #     sheets = list(wb_lake_file.sheetnames)
    #     if sheets[0].upper() == "COVER":
    #         del sheets[0]
    #     for sheet1 in sheets:
    #         sheet = wb_lake_file[sheet1]
    #         # tgt_mpng = sheet.cell(row=3, column=2).value
    #         # tgt_workflow = sheet.cell(row=4, column=2).value
    #         # src_db = sheet.cell(row=6, column=2).value
    #         tgt_db = sheet.cell(row=9, column=18).value
    #         # src_tb_name = sheet.cell(row=11, column=1).value
    #         tgt_tb_name = sheet.cell(row=9, column=19).value
    #         # src_fltr = sheet.cell(row=11, column=11).value
    #         # src_join = sheet.cell(row=11, column=12).value
    #         # tgt_fltr = sheet.cell(row=11, column=13).value
    #         # tgt_join = sheet.cell(row=11, column=14).value
    #         # tgt_data_typs = []
    #         tgt_cols = []
    #         # trnsfrmtns = []
    #         tgt_data_typs = []
    #         for i in range(9, sheet.max_row + 1):
    #             temp_tgt_col = sheet.cell(row=i, column=20).value
    #             temp_tgt_dttype = sheet.cell(row=i, column=21).value
    #             if temp_tgt_col is not None and temp_tgt_col != 'NA':
    #                 tgt_cols.append(temp_tgt_col)
    #             if temp_tgt_dttype is not None and temp_tgt_col != 'NA':
    #                 tgt_data_typs.append(temp_tgt_dttype)
    #
    #         # appending the data into list in the format of Src_cols,Tgt_cols,src_db,tgt_db,src_tb_name,tgt_tb_name,wf,mp,trnfromations,srcfltrand join, tgt flr an djoj
    #         sttm_data.append([tgt_db, tgt_tb_name, tgt_cols, tgt_data_typs])
    #
    #     return sttm_data
    #     # Calling the Execute function in Begin class
    #     # Begin.Execute('', self.src_sys_name, sttm_data, self.queries_data, self.exe_data, input_file_path)

# filename = r'C:\Users\MadhReddy\Desktop\AWS Testing Accelerator - Copy\Input Files\STTM_DAB_HIVE_Lake_LakeHist - Copy.xlsx'
# template = r'C:\Users\MadhReddy\Desktop\AWS Testing Accelerator - Copy\Pyfiles\STTM_Lake_Mart_Template.xlsx'
# validate = ValidateSTTM('src_sys_name', r'C:\Users\MadhReddy\Desktop',template)
# validate.validate_lake_mart_sttm(filename)