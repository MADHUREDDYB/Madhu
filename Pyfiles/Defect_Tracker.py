import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import openpyxl
from jira import JIRA
from openpyxl import Workbook
import datetime
import os

class DefectTracker:

    def __init__(self, src_sys_name, output_folder_nm, run_time):
        self.src_sys_name = src_sys_name
        self.output_folder_nm = output_folder_nm
        self.run_time = run_time

    def createdefecttracker(self,prj_key,reporter,assignee):
        folder = self.output_folder_nm + '\\Output Files\\' + 'Defect Tracker'
        if not os.path.exists(folder):
            os.makedirs(folder)
        wb = Workbook()
        ws1 = wb.create_sheet(title=self.src_sys_name)
        headers = ['Project Key','Summary', 'Description', 'Priority', 'Environment', 'Issue Type', 'Reporter', 'Assignee','JIRA Defect ID']
        ws1.freeze_panes = 'B2'
        ws1.sheet_view.zoomScale = 70
        ws1.auto_filter.ref = 'A1:I1'
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 80
        ws1.column_dimensions['C'].width = 80
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 20
        ws1.column_dimensions['H'].width = 20
        ws1.append(headers)
        for m in range(1, len(headers) + 1):
            ws1.cell(row=1, column=m).fill = PatternFill(start_color="575757", end_color="575757",
                                                         fill_type="solid")
            ws1.cell(row=1, column=m).font = Font(color='6FF242')

        src_lake_path = self.output_folder_nm + '\\output files\\' + 'Src-Lake_Validations'
        metadatavalidation_path = self.output_folder_nm + '\\output files\\' + 'MetadaValidations'
        lake_mart_path = self.output_folder_nm + '\\output files\\' + 'ETL Validations'
        defects = []

        if not os.path.exists(src_lake_path):
            sub_fldr = os.listdir(src_lake_path)
            print(sub_fldr)
            for sub_sub_fldr in sub_fldr:
                temp_src_lake_path = src_lake_path + '\\' + sub_sub_fldr
                all_files = os.listdir(temp_src_lake_path)
                for file in all_files:
                    file_path = temp_src_lake_path + '\\' + file
                    temp_wb = openpyxl.load_workbook(file_path)
                    for sheet in temp_wb.worksheets:
                        color = sheet.sheet_properties.tabColor
                        if color is not None:
                            if color.rgb == '00FE635E':
                                print('create defect')
                                summary = 'SIT- %s - %s - Src DataValidation' % (file.split('-')[0], sub_sub_fldr)
                                desc = 'SIT -%s- Src DataValidation failed in %s table' % (
                                sub_sub_fldr, file.split('-')[0])
                                defects.append([prj_key, summary, desc, 'Medium', 'SIT', 'Defect', reporter, assignee])


        if not os.path.exists(metadatavalidation_path):
            sub_fldr = os.listdir(metadatavalidation_path)
            print(sub_fldr)
            for sub_sub_fldr in sub_fldr:
                temp_metadatavalidation_path = metadatavalidation_path + '\\' + sub_sub_fldr
                all_files = os.listdir(temp_metadatavalidation_path)
                all_files.reverse()
                file = all_files[0]
                file_path = temp_metadatavalidation_path + '\\' + file
                temp_wb = openpyxl.load_workbook(file_path)
                summary_sheet = temp_wb['Summary']
                if summary_sheet.cell(row=2, column=3).value == 'Validation Successful':
                    pass
                else:
                    for row_num in range(2, summary_sheet.max_row + 1):
                        if summary_sheet.cell(row=row_num, column=2).value == 'Fail':
                            table = summary_sheet.cell(row=row_num, column=1).value
                            summary = 'SIT- %s - %s - MetaDataValidation' % (table, sub_sub_fldr)
                            desc = 'SIT - Layer : %s - MetaDataValidation failed in %s table' % (sub_sub_fldr, table)
                            defects.append([prj_key, summary, desc, 'Medium', 'SIT', 'Defect', reporter, assignee])

        if not os.path.exists(lake_mart_path):
            sub_fldr = os.listdir(lake_mart_path)
            print(sub_fldr)
            for sub_sub_fldr in sub_fldr:
                temp_lake_mart_path = lake_mart_path + '\\' + sub_sub_fldr
                all_files = os.listdir(temp_lake_mart_path)
                all_files.reverse()
                file = all_files[0]
                print(file)
                file_path = temp_lake_mart_path + '\\' + file
                temp_wb = openpyxl.load_workbook(file_path)
                for sheet in temp_wb.worksheets:
                    for row_num in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row_num, column=13).value == 'Fail':
                            table = sheet.cell(row=row_num, column=3).value
                            level = sheet.cell(row=row_num, column=5).value
                            column_nm = sheet.cell(row=row_num, column=4).value
                            tc_name = sheet.cell(row=row_num, column=6).value
                            exp_val = sheet.cell(row=row_num, column=11).value
                            act_val = sheet.cell(row=row_num, column=12).value
                            if level == 'Table Level':
                                summary = 'SIT- %s - %s - %s - %s' % (sub_sub_fldr, level, tc_name, table)
                                desc = 'SIT - Layer : %s - %s - %s failed in %s table - Expected value:%s - Actual Value: %s' % (
                                sub_sub_fldr, level, tc_name, table, exp_val, act_val)
                            else:
                                summary = 'SIT- %s - %s - %s - %s' % (sub_sub_fldr, level, tc_name, table)
                                desc = 'SIT - Layer : %s - %s - %s column - %s failed in %s table - Expected value:%s - Actual Value: %s' % (
                                sub_sub_fldr, level, column_nm, tc_name, table, exp_val, act_val)
                            defects.append([prj_key, summary, desc, 'Medium', 'SIT', 'Defect', reporter, assignee])

        # defects = list(set(defects))
        for defect in defects:
            ws1.append(defect)
        wb.remove(wb['Sheet'])
        wb.save(folder + '\\%s_%s.xlsx' % (self.src_sys_name, self.run_time))

    def createdefects(self, server,API_token,username):
        print(API_token,username)
        folder = self.output_folder_nm + '\\Output Files\\Defect Tracker'
        if not os.path.exists(folder):
            return '<b>EXECUTION STOPPED! : </b>Folder not found %s: ' % folder
        filename = self.output_folder_nm + '\\Output Files\\Defect Tracker'
        # list_of_files = os.listdir(filename)  # * means all if need specific format then *.csv
        files = os.listdir(filename)
        if not files:
            return '<b>EXECUTION STOPPED! : </b>The Defect Tracker is not present at %s: ' % filename
        files = files[::-1][0]
        filename = filename + '\\' + files
        wb = openpyxl.load_workbook(filename)

        jira = JIRA(basic_auth=(username, API_token),
                    options={'server': server})
        defectid = []
        for sheet in wb.worksheets:
            for rownum in range(2, sheet.max_row + 1):
                # temp =
                temp = (jira.create_issue(fields={
                    'project': {'key': sheet.cell(row=rownum, column=1).value},
                    'summary': sheet.cell(row=rownum, column=2).value,
                    'description': sheet.cell(row=rownum, column=3).value,
                    'priority': {"name": sheet.cell(row=rownum, column=4).value},
                    'environment': sheet.cell(row=rownum, column=5).value,
                    'issuetype': {"name": sheet.cell(row=rownum, column=6).value},
                    # #  'reporter': 'farhhan.adil',
                    # 'assignee': {"name": "farhhan adil", 'email':'farhhan.adil12@gmail.com'},
                }))
                sheet.cell(row=rownum,column=9).value = temp.key

        wb.save(filename)





